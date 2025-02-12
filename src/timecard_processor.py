import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import load_workbook
from tqdm import tqdm

class TimecardProcessor:
    def __init__(self):
        self.exception_codes = {
            '00': '通常', '01': '早出', '02': '遅刻', '03': '外出',
            '04': '再入', '05': '早退', '06': '残業', '07': '徹夜',
            '09': '深夜', '10': '休日出勤', '11': '休日早出',
            '12': '休日遅刻', '13': '休日外出', '14': '休日再入',
            '15': '休日早退', '16': '休日残業', '17': '休日徹夜',
            '19': '休日深夜'
        }
        self.name_mapping = {}

    def read_csv_with_encoding(self, file_path, encodings=['utf-8', 'cp932', 'shift_jis', 'shift_jisx0213']):
        """複数のエンコーディングでCSVファイルの読み込みを試行"""
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, 
                                encoding=encoding,
                                skiprows=1, 
                                header=None,
                                names=['カード番号', '区分', '年月日',
                                      '入1時刻', '入1異例', '退1時刻', '退1異例',
                                      '入2時刻', '入2異例', '退2時刻', '退2異例',
                                      '時数1', '時数2', '空白'],
                                skipinitialspace=True)
                print(f"ファイルを読み込みました（エンコーディング: {encoding}）")
                return df
            except UnicodeDecodeError:
                continue
        raise ValueError("サポートされているエンコーディングでファイルを読み込めませんでした")

    def process_csv(self, input_file, output_file, mapping_file):
        """CSVファイルを処理してExcelファイルに出力"""
        # CSVファイルを読み込み
        df = self.read_csv_with_encoding(input_file)
        
        # マッピングファイルを読み込み（必要な場合）
        if mapping_file:
            mapping_df = self.read_csv_with_encoding(mapping_file)
            card_col = mapping_df.columns[0]
            name_col = mapping_df.columns[1]
            
            # カード番号を文字列に変換し、4桁にゼロ埋め
            mapping_df[card_col] = mapping_df[card_col].astype(str).str.zfill(4)
            self.name_mapping = dict(zip(mapping_df[card_col], mapping_df[name_col]))
            
            # カード番号を4桁にゼロ埋め
            df['カード番号'] = df['カード番号'].astype(str).str.zfill(4)
            
            # 名前変換を実行
            df['名前'] = df['カード番号'].map(self.name_mapping)

        # 不要な列を削除
        df = df.drop('空白', axis=1)

        # データのクリーニング
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].str.strip()

        # カード番号を名前に変換（マッピングがある場合）
        if self.name_mapping:
            df['カード番号'] = df['カード番号'].astype(str).str.zfill(4)
            df['名前'] = df['カード番号'].map(self.name_mapping)
            df['名前'] = df.apply(
                lambda row: row['名前'] if pd.notna(row['名前']) 
                else f"未登録(カード番号:{row['カード番号']})", axis=1
            )
            df = df.drop('カード番号', axis=1)
            df = df.rename(columns={'名前': 'カード番号'})

        # 年月日を日付のみに変換（時間情報なし）
        df['年月日'] = pd.to_datetime(df['年月日'].astype(str), format='%y/%m/%d').dt.date
        
        # 時刻データを時刻文字列として保持
        time_columns = ['入1時刻', '退1時刻', '入2時刻', '退2時刻']
        for col in time_columns:
            df[col] = df[col].apply(lambda x: x if pd.notna(x) and str(x).strip() else '')

        # 時数を文字列として保持し、合計を計算
        def format_duration(duration_str):
            if pd.isna(duration_str) or str(duration_str).strip() == '':
                return ''
            try:
                hours, minutes = map(int, str(duration_str).split(':'))
                return f'{hours:03d}:{minutes:02d}'
            except:
                return ''

        duration_columns = ['時数1', '時数2']
        for col in duration_columns:
            df[col] = df[col].apply(format_duration)

        # 異例コードを変換
        exception_columns = ['入1異例', '退1異例', '入2異例', '退2異例']
        for col in exception_columns:
            df[col] = df[col].astype(str).map(self.exception_codes).fillna('')

        # 時数1と時数2の合計を計算
        def convert_duration_to_minutes(duration_str):
            if not duration_str or str(duration_str).strip() == '':
                return 0
            try:
                hours, minutes = map(int, duration_str.split(':'))
                return hours * 60 + minutes
            except:
                return 0

        def convert_minutes_to_duration(minutes):
            if minutes == 0:
                return ''
            hours = minutes // 60
            remaining_minutes = minutes % 60
            return f'{hours:03d}:{remaining_minutes:02d}'

        df['合計時数'] = df.apply(
            lambda row: convert_minutes_to_duration(
                convert_duration_to_minutes(row['時数1']) +
                convert_duration_to_minutes(row['時数2'])
            ),
            axis=1
        )

        # Excelファイルとして出力
        print("Excelファイルに出力中...")
        df.to_excel(output_file, sheet_name='勤怠データ', index=False)

        # Excelファイルを開いてフォーマットを設定
        wb = load_workbook(output_file)
        ws = wb['勤怠データ']

        # 列のインデックスを取得
        date_col_idx = df.columns.get_loc('年月日') + 1
        time_col_indices = [df.columns.get_loc(col) + 1 for col in time_columns]

        # セルの書式設定
        for row in range(2, ws.max_row + 1):  # ヘッダーを除く
            # 日付書式の設定
            ws.cell(row=row, column=date_col_idx).number_format = 'YYYY/MM/DD'
            
            # 時刻書式の設定
            for col_idx in time_col_indices:
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:  # 値が存在する場合のみ
                    cell.number_format = 'HH:MM'

        # 変更を保存
        wb.save(output_file)
        print("処理が完了しました")
        return df